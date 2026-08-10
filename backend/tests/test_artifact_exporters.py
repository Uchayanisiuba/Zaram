"""The exporters: what comes out, and what happens when a format cannot run.

Two things are being defended here, and only one of them is "the file is valid".

**The acceptance criterion is a link that works in Word.** M9b asks for "a .docx
where claims link back to the source paragraph they came from", and that has to
hold with Zaram not running — a citation that only resolves inside the app that
made it is not something you can send to a client. So the .docx tests unzip the
file and assert that every hyperlink anchor resolves to a bookmark that is
actually present. Word does not report a broken anchor; it renders it as text
that looks like a link and goes nowhere, which is a failure no smoke test would
catch by looking at the document.

**Unavailability is an answer, not an exception.** PDF cannot run on a Windows
box without the GTK runtime, and the way that fact reaches the user is the
difference between "Zaram is broken" and "Zaram needs a thing installed". These
tests assert the *shape* of that answer rather than its content, so they pass on
a machine where PDF does work.
"""

from __future__ import annotations

import ast
import re
import zipfile
from io import BytesIO
from pathlib import Path

import pytest

from artifacts import export
from artifacts.contracts import Artifact, ArtifactSource, Claim
from artifacts.export import _reader
from artifacts.export.base import Availability, ExportUnavailable
from artifacts.export.chart import PALETTE, ChartSpec, Series, TooManySeries, build_png
from artifacts.export.docx import _bookmark_name
from artifacts.html import render_chart, render_document, render_spreadsheet
from artifacts.store import ArtifactStore

EXPORT_DIR = Path(__file__).resolve().parents[1] / "artifacts" / "export"


CLAIMS = [
    Claim(
        id="c1",
        source_id="memory:55b6",
        excerpt="Northwind pay on 30-day terms.",
        source_excerpt="Clause 4.2: payment due within thirty (30) days.",
        source_revision="rev-3",
    ),
    Claim(
        id="c2",
        source_id="doc:brief",
        excerpt="The agreed day rate is 85,000 naira.",
        source_excerpt="Day rate: NGN 85,000.",
    ),
]

SOURCES = [
    ArtifactSource(kind="document", title="Master agreement", url="file:///n.pdf"),
    ArtifactSource(kind="memory", title="Rate agreed in April"),
]


@pytest.fixture
def document_html() -> str:
    return render_document(
        title="Proposal — Northwind",
        blocks=["Here is the scope.", CLAIMS[0], CLAIMS[1], "Work begins on signature."],
        sources=SOURCES,
        claims=CLAIMS,
    )


@pytest.fixture
def spreadsheet_html() -> str:
    return render_spreadsheet(
        title="Invoices",
        caption="Northwind",
        header=["Invoice", "Issued", "Amount", "Share"],
        rows=[
            ["INV-041", "2026-07-02", "₦425,000", "50%"],
            ["INV-042", "2026-07-19", "1,275,000", "12.5%"],
        ],
        sources=SOURCES,
    )


class TestAvailabilityIsAnAnswer:
    """CLAUDE.md: "Disabled capabilities are visible, not silent.\""""

    def test_every_exporter_can_be_asked_without_raising(self):
        """The check itself must never be the thing that fails — it runs on
        every render of the format picker."""
        for extension, exporter in export.EXPORTERS.items():
            availability = exporter.availability()
            assert isinstance(availability, Availability), extension

    def test_unavailable_formats_are_still_listed(self):
        """Silently dropping them would leave a user wondering whether Zaram
        can do PDF at all, rather than knowing it needs something installed."""
        listed = {extension for extension, _ in export.formats()}

        assert listed == set(export.EXPORTERS)

    def test_available_formats_sort_first(self):
        availability_flags = [availability.ok for _, availability in export.formats()]

        assert availability_flags == sorted(availability_flags, reverse=True)

    def test_an_unavailable_format_states_a_reason_and_a_remedy(self):
        for extension, availability in export.formats():
            if availability.ok:
                continue
            assert availability.reason, f"{extension} is off with no reason given"
            assert availability.remedy, f"{extension} says no with nothing to do"

    def test_markdown_is_available_everywhere(self):
        """The floor. Whatever else is missing, the user can still get a file
        out, which is what keeps a packaging gap from being a dead end."""
        assert export.get("md").availability().ok

    def test_rendering_an_unavailable_format_raises_with_the_reason(
        self, monkeypatch, document_html
    ):
        blocked = Availability(ok=False, reason="needs a thing", remedy="install it")
        monkeypatch.setattr(export.get("pdf"), "availability", lambda: blocked)

        with pytest.raises(ExportUnavailable) as caught:
            export.render(document_html, "pdf")

        # Not an ImportError, and not a bare traceback naming a DLL: a caller
        # that skipped the check still gets something it can show a user.
        assert "needs a thing" in str(caught.value)
        assert "install it" in str(caught.value)

    def test_an_unknown_format_is_a_different_failure_from_an_unavailable_one(self):
        with pytest.raises(export.UnknownFormat):
            export.get("psd")


class TestMarkdown:
    def test_claims_become_footnotes(self, document_html):
        text = export.render(document_html, "md").decode("utf-8")

        assert "Northwind pay on 30-day terms.[^1]" in text
        assert "The agreed day rate is 85,000 naira.[^2]" in text

    def test_every_footnote_reference_resolves(self, document_html):
        text = export.render(document_html, "md").decode("utf-8")

        referenced = set(re.findall(r"\[\^(\d+)\](?!:)", text))
        defined = set(re.findall(r"^\[\^(\d+)\]:", text, flags=re.MULTILINE))

        assert referenced == defined, "a footnote marker points at nothing"

    def test_the_footnote_carries_the_source_excerpt(self, document_html):
        text = export.render(document_html, "md").decode("utf-8")

        assert "Clause 4.2" in text.split("## Claims")[1]

    def test_it_is_utf8_and_keeps_the_real_characters(self, document_html):
        text = export.render(document_html, "md").decode("utf-8")

        assert "—" in text, "the em dash was mangled"


class TestDocxClaimsLinkToTheirSource:
    """M9b's acceptance criterion, asserted against the file rather than the API."""

    @staticmethod
    def _document_xml(data: bytes) -> str:
        with zipfile.ZipFile(BytesIO(data)) as archive:
            return archive.read("word/document.xml").decode("utf-8")

    def test_every_link_resolves_to_a_bookmark_that_exists(self, document_html):
        """Word does not complain about a dangling anchor. It draws a link that
        does nothing, which reads as working provenance and is not."""
        xml = self._document_xml(export.render(document_html, "docx"))

        anchors = re.findall(r'w:anchor="([^"]+)"', xml)
        bookmarks = set(re.findall(r'w:bookmarkStart[^>]*w:name="([^"]+)"', xml))

        assert anchors, "no claim was linked at all"
        assert set(anchors) <= bookmarks, (
            f"these anchors point at nothing: {set(anchors) - bookmarks}"
        )

    def test_there_is_one_link_per_claim(self, document_html):
        xml = self._document_xml(export.render(document_html, "docx"))

        assert len(re.findall(r'w:anchor="', xml)) == len(CLAIMS)

    def test_the_source_excerpt_travels_into_the_file(self, document_html):
        """What makes the document defensible away from Zaram: the reader can
        compare the sentence against the source without the app."""
        xml = self._document_xml(export.render(document_html, "docx"))

        assert "Clause 4.2" in xml

    @pytest.mark.parametrize(
        "anchor",
        ["claim-c1", "claim-a-b-c", "claim-" + "x" * 80, "1-leading-digit", "-"],
    )
    def test_bookmark_names_are_ones_word_will_accept(self, anchor):
        """Letters, digits and underscore; must start with a letter; 40 max.
        Word drops a name breaking these silently, taking every link with it."""
        name = _bookmark_name(anchor)

        assert name, "produced an empty bookmark name"
        assert len(name) <= 40
        assert name[0].isalpha()
        assert all(character.isalnum() or character == "_" for character in name)

    def test_a_claim_with_no_entry_in_sources_is_not_linked_into_air(self):
        """Better an unlinked sentence than a link to a bookmark that was never
        written — the second looks like working provenance."""
        orphan = Claim(id="c9", source_id="memory:x", excerpt="Unsupported.")
        html = render_document(title="T", blocks=[orphan], sources=[], claims=[])

        xml = self._document_xml(export.render(html, "docx"))

        anchors = re.findall(r'w:anchor="([^"]+)"', xml)
        bookmarks = set(re.findall(r'w:bookmarkStart[^>]*w:name="([^"]+)"', xml))
        assert set(anchors) <= bookmarks

    def test_the_document_says_when_nothing_was_recalled(self):
        """A real state, stated. Not an empty Sources heading, which reads as a
        rendering failure."""
        html = render_document(title="T", blocks=["Plain prose."], sources=[], claims=[])

        xml = self._document_xml(export.render(html, "docx"))

        assert "Nothing was recalled" in xml


class TestSpreadsheet:
    @staticmethod
    def _load(data: bytes):
        from openpyxl import load_workbook

        return load_workbook(BytesIO(data))

    def test_currency_and_separators_become_a_real_number(self, spreadsheet_html):
        sheet = self._load(export.render(spreadsheet_html, "xlsx")).worksheets[0]

        assert sheet["C2"].value == 425000
        assert sheet["C3"].value == 1275000

    def test_percentages_stay_text(self, spreadsheet_html):
        """"50%" is 0.5 to Excel and 50 to a naive strip. Text is visibly
        unfinished; the wrong number is invisibly wrong."""
        sheet = self._load(export.render(spreadsheet_html, "xlsx")).worksheets[0]

        assert sheet["D2"].value == "50%"

    def test_dates_stay_text(self, spreadsheet_html):
        sheet = self._load(export.render(spreadsheet_html, "xlsx")).worksheets[0]

        assert sheet["B2"].value == "2026-07-02"

    def test_the_filter_spans_the_data_and_not_just_the_header(self, spreadsheet_html):
        """A filter of `A1:D1` opens in Excel with working dropdowns that match
        nothing — correct-looking and silently useless, in a file the user
        sends to someone else."""
        sheet = self._load(export.render(spreadsheet_html, "xlsx")).worksheets[0]

        assert sheet.auto_filter.ref is not None
        last_row = int(re.search(r"(\d+)$", sheet.auto_filter.ref).group(1))
        assert last_row == sheet.max_row > 1

    def test_the_header_is_frozen(self, spreadsheet_html):
        sheet = self._load(export.render(spreadsheet_html, "xlsx")).worksheets[0]

        assert sheet.freeze_panes == "A2"

    def test_sources_get_their_own_sheet(self, spreadsheet_html):
        """Under the data they would land inside the filter range and sort with
        the rows — provenance corrupting the table it documents."""
        workbook = self._load(export.render(spreadsheet_html, "xlsx"))

        assert "Sources" in workbook.sheetnames
        assert workbook["Sources"]["A1"].value == "Sources"
        assert "Master agreement" in str(workbook["Sources"]["A2"].value)

    def test_a_headerless_table_is_not_frozen(self):
        """Freezing unconditionally locks a data row off the top of the sheet."""
        html = render_spreadsheet(title="T", header=[], rows=[["a", "b"]])

        sheet = self._load(export.render(html, "xlsx")).worksheets[0]

        assert sheet.freeze_panes is None

    def test_a_caption_with_illegal_characters_does_not_fail_the_export(self):
        """openpyxl raises on a bad sheet name. Losing a whole export because a
        table caption contained a colon is not a trade worth making."""
        html = render_spreadsheet(
            title="T", caption="Q3: profit/loss [draft]", header=["a"], rows=[["1"]]
        )

        workbook = self._load(export.render(html, "xlsx"))

        assert all(len(name) <= 31 for name in workbook.sheetnames)


class TestCharts:
    SPEC = ChartSpec(
        title="Revenue",
        kind="bar",
        categories=["Apr", "May"],
        series=[Series("Northwind", [425000, 1275000])],
    )

    def test_it_draws_a_png(self):
        png = build_png(self.SPEC)

        assert png[:8] == b"\x89PNG\r\n\x1a\n"

    @pytest.mark.parametrize("kind", ["bar", "hbar", "line"])
    def test_every_form_renders(self, kind):
        spec = ChartSpec(
            title="T", kind=kind, categories=["a", "b"],
            series=[Series("s", [1, 2]), Series("t", [2, 1])],
        )

        assert build_png(spec)[:8] == b"\x89PNG\r\n\x1a\n"

    def test_a_ninth_series_is_refused_rather_than_given_an_invented_hue(self):
        """A generated colour is one the reader cannot distinguish and the
        caller cannot predict. Facet, or fold the tail into a total."""
        spec = ChartSpec(
            title="T", kind="bar", categories=["a"],
            series=[Series(str(n), [n]) for n in range(len(PALETTE) + 1)],
        )

        with pytest.raises(TooManySeries):
            build_png(spec)

    def test_colours_are_assigned_in_fixed_order_and_never_cycled(self):
        assert len(set(PALETTE)) == len(PALETTE)

    def test_an_empty_chart_is_refused(self):
        with pytest.raises(ValueError):
            build_png(ChartSpec(title="T", kind="bar", categories=[], series=[]))

    def test_an_unknown_kind_is_refused(self):
        spec = ChartSpec(title="T", kind="donut", categories=["a"],
                         series=[Series("s", [1])])

        with pytest.raises(ValueError):
            build_png(spec)

    def test_the_png_round_trips_through_the_html(self):
        """HTML is the source of truth, so the image has to survive being
        embedded in it and pulled back out byte-for-byte."""
        png = build_png(self.SPEC)
        html = render_chart(title="Revenue", png=png, header=["Month"], rows=[["Apr"]])

        assert export.render(html, "png") == png

    def test_the_numbers_are_always_under_the_picture(self):
        """The relief the palette's low-contrast slots require, and the thing
        that makes a chart checkable rather than merely decorative."""
        png = build_png(self.SPEC)
        html = render_chart(
            title="Revenue", png=png,
            header=["Month", "Amount"], rows=[["Apr", 425000]],
        )

        assert "425000" in html
        assert "<table>" in html

    def test_exporting_a_document_as_png_says_what_is_wrong(self, document_html):
        with pytest.raises(ValueError, match="does not contain one"):
            export.render(document_html, "png")


class TestTheReader:
    def test_claims_keep_their_ids_through_the_parse(self, document_html):
        document = _reader.read(document_html)

        claim_ids = {
            run.claim_id
            for block in document.body_blocks()
            for run in block.runs
            if run.claim_id
        }
        assert claim_ids == {"c1", "c2"}

    def test_the_sources_section_is_distinguished_from_the_body(self, document_html):
        document = _reader.read(document_html)

        assert document.body_blocks()
        assert document.source_blocks()
        assert not any(block.in_sources for block in document.body_blocks())

    def test_the_title_survives(self, document_html):
        assert _reader.read(document_html).title == "Proposal — Northwind"

    def test_stylesheet_text_is_not_read_as_prose(self, document_html):
        """The `<style>` block is inline, so a naive walk puts CSS in the .docx."""
        document = _reader.read(document_html)

        assert "border-collapse" not in " ".join(b.text for b in document.blocks)
        assert "font:" not in " ".join(b.text for b in document.blocks)

    def test_a_header_row_is_distinguished_from_a_data_row(self, spreadsheet_html):
        table = _reader.read(spreadsheet_html).tables[0]

        assert table.header == ["Invoice", "Issued", "Amount", "Share"]
        assert len(table.rows) == 2

    def test_escaped_characters_come_back_decoded_once(self):
        html = render_document(title="A & B", blocks=["x < y & z"], sources=[], claims=[])

        document = _reader.read(html)

        assert document.title == "A & B"
        assert any("x < y & z" in block.text for block in document.body_blocks())


class TestWritingThroughTheStore:
    def test_the_record_describes_the_file_that_exists(self, tmp_path, document_html):
        store = ArtifactStore(tmp_path)
        artifact = Artifact(filename="proposal", html=document_html)

        path = export.write(artifact, "docx", store)

        assert Path(artifact.path) == path
        assert artifact.filename == path.name
        assert artifact.size_bytes == path.stat().st_size

    def test_the_extension_is_added_when_the_caller_omits_it(self, tmp_path, document_html):
        store = ArtifactStore(tmp_path)

        path = export.write(Artifact(html=document_html), "md", store, filename="notes")

        assert path.name == "notes.md"

    def test_a_second_export_never_replaces_the_first(self, tmp_path, document_html):
        """The record has to follow the name the store actually used, not the
        one that was asked for."""
        store = ArtifactStore(tmp_path)

        first = export.write(Artifact(html=document_html), "md", store, filename="p")
        second_artifact = Artifact(html=document_html)
        second = export.write(second_artifact, "md", store, filename="p")

        assert first != second
        assert second.name == "p-2.md"
        assert second_artifact.filename == "p-2.md"

    def test_a_model_proposed_traversal_cannot_escape(self, tmp_path, document_html):
        store = ArtifactStore(tmp_path / "out")

        path = export.write(
            Artifact(html=document_html), "md", store, filename="../../escape"
        )

        assert path.parent == store.root


class TestTheExportersDoNotTouchTheFilesystem:
    """`ArtifactStore` is the only module that creates files, and it has no
    capability to remove or replace one. That guarantee is only worth having if
    adding a sixth format cannot quietly route around it — so this scans the
    export package the same way `test_artifact_write_path.py` scans the store.

    It matches *calls*, not names: `workbook.save(buffer)` and
    `HTML(...).write_pdf()` both write to memory, and a name-based scan would
    fail them while missing an `open()` that actually hits the disk.
    """

    FORBIDDEN = {
        ("os", "remove"), ("os", "unlink"), ("os", "rmdir"), ("os", "replace"),
        ("os", "rename"), ("os", "makedirs"), ("os", "mkdir"),
        ("shutil", "rmtree"), ("shutil", "move"), ("shutil", "copyfile"),
        ("Path", "unlink"), ("Path", "rmdir"), ("Path", "replace"),
        ("Path", "rename"), ("Path", "write_bytes"), ("Path", "write_text"),
        ("Path", "mkdir"), ("Path", "open"),
    }

    @staticmethod
    def _dotted(node: ast.AST) -> tuple:
        parts = []
        while isinstance(node, ast.Attribute):
            parts.append(node.attr)
            node = node.value
        if isinstance(node, ast.Name):
            parts.append(node.id)
        return tuple(reversed(parts))

    @pytest.mark.parametrize(
        "module", sorted(path.name for path in EXPORT_DIR.glob("*.py"))
    )
    def test_no_exporter_opens_or_unmakes_a_file(self, module):
        tree = ast.parse((EXPORT_DIR / module).read_text(encoding="utf-8"))
        offenders = []

        for node in ast.walk(tree):
            if not isinstance(node, ast.Call):
                continue
            dotted = self._dotted(node.func)
            if not dotted:
                continue
            if dotted[-1] == "open" and len(dotted) == 1:
                offenders.append(f"line {node.lineno}: open(...)")
            if len(dotted) >= 2 and dotted[-2:] in self.FORBIDDEN:
                offenders.append(f"line {node.lineno}: {'.'.join(dotted)}")

        assert not offenders, (
            f"{module} touches the filesystem:\n  " + "\n  ".join(offenders)
            + "\n\nExporters return bytes. ArtifactStore is the only writer."
        )
