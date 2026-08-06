"""HTML tables → .xlsx.

The tables in the HTML are the spreadsheet. That is the whole design, and it is
the reason a spreadsheet goes through HTML at all rather than straight from data
to openpyxl: the preview a user is shown before the file is written is the same
table the exporter reads, so the two cannot disagree.

Numbers are written as numbers
------------------------------
A figure that arrives in Excel as text is a spreadsheet that will not sum, and a
user discovers that in front of a client rather than here. So cells that parse
as a number are written as one, with currency symbols and thousands separators
removed first — those are formatting the HTML applied, not part of the value.

The conversion is deliberately conservative, and the cases it *refuses* are the
considered part:

- **Percentages stay text.** "50%" is 0.5 to Excel and 50 to a naive strip, and
  writing the wrong one of those into a cell that then feeds a formula is the
  exact failure this module is supposed to prevent. Text is visibly unfinished;
  a wrong number is invisibly wrong.
- **Dates stay text.** Excel's willingness to guess at date-like strings is a
  well-known way to corrupt data, and it guesses differently by locale.

"1,234" is 1234 everywhere, so that one converts.
"""

from __future__ import annotations

import io
import re
from typing import Union

from . import _reader
from .base import Availability, module_available

#: Leading currency symbols, which are formatting rather than value. Trailing
#: percent is deliberately *not* here — see the module docstring.
_CURRENCY = re.compile(r"^[\s$£€₦¥]+")
_NUMERIC = re.compile(r"^-?\d{1,3}(,\d{3})*(\.\d+)?$|^-?\d+(\.\d+)?$")

#: Excel refuses sheet names over 31 characters or containing these.
_SHEET_ILLEGAL = re.compile(r"[\\/*?:\[\]]")


class XlsxExporter:
    extension = "xlsx"
    media_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    label = "Excel spreadsheet"

    def availability(self) -> Availability:
        return module_available("openpyxl", needed_for="Excel export")

    def export(self, document_html: str, *, filename: str = "") -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter

        doc = _reader.read(document_html)
        workbook = Workbook()
        workbook.remove(workbook.active)

        tables = doc.tables or [_reader.Table(caption=doc.title)]

        for index, table in enumerate(tables, start=1):
            sheet = workbook.create_sheet(
                _sheet_name(table.caption or doc.title, index, len(tables))
            )

            if table.header:
                sheet.append(list(table.header))
                for cell in sheet[1]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(vertical="center")

            for row in table.rows:
                sheet.append([_coerce(cell) for cell in row])

            if table.header:
                # Freeze and filter only when there is a header — doing it
                # unconditionally locks a data row off the top of a headerless
                # sheet.
                #
                # After the rows, not before: `sheet.dimensions` describes what
                # is in the sheet *now*, so setting the filter first produces
                # `A1:D1` — a filter over the header alone, which opens in Excel
                # with working dropdowns that match nothing. It looks correct
                # and silently filters no data, which is the worst shape for a
                # bug in a file the user sends to someone else.
                sheet.freeze_panes = "A2"
                sheet.auto_filter.ref = sheet.dimensions

            _fit_columns(sheet, get_column_letter)

        _write_sources_sheet(workbook, doc)

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()


def _coerce(value: str) -> Union[str, int, float]:
    """A cell's text as a number when it unambiguously is one, else as text."""
    stripped = value.strip()
    if not stripped:
        return ""

    candidate = _CURRENCY.sub("", stripped).replace(",", "")
    if not _NUMERIC.match(_CURRENCY.sub("", stripped)):
        return value

    try:
        return int(candidate) if "." not in candidate else float(candidate)
    except ValueError:
        return value


def _sheet_name(proposed: str, index: int, total: int) -> str:
    """A tab name Excel will accept.

    Excel does not report a rejected sheet name — openpyxl raises, and a whole
    export failing because a table caption contained a colon is not a trade
    worth making.
    """
    name = _SHEET_ILLEGAL.sub("-", proposed or "").strip() or "Data"
    if total > 1:
        suffix = f" ({index})"
        name = f"{name[: 31 - len(suffix)]}{suffix}"
    return name[:31]


def _fit_columns(sheet, get_column_letter) -> None:
    """Width from the longest cell, capped.

    Uncapped, one long source excerpt makes a column wider than the screen and
    the sheet unreadable. Excel has no autofit that can be triggered from a
    file, so this is an approximation done once at write time.
    """
    for column_index, column in enumerate(sheet.columns, start=1):
        longest = max((len(str(cell.value or "")) for cell in column), default=0)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(longest + 2, 9), 60
        )


def _write_sources_sheet(workbook, doc: _reader.Document) -> None:
    """Provenance as its own sheet.

    A spreadsheet has nowhere to put a footnote, and a Sources block pasted
    under the data would be read as data — it would land inside the auto-filter
    range and sort with the rows. Its own sheet is the only placement that does
    not corrupt the table it documents.
    """
    from openpyxl.styles import Alignment, Font

    entries = [
        block.text.strip()
        for block in doc.source_blocks()
        if block.tag == "li" and block.text.strip()
    ]

    sheet = workbook.create_sheet("Sources")
    sheet.column_dimensions["A"].width = 100
    heading = sheet.cell(row=1, column=1, value="Sources")
    heading.font = Font(bold=True)

    if not entries:
        sheet.cell(
            row=2,
            column=1,
            value=(
                "Nothing was recalled for this spreadsheet. It was built from "
                "what you provided, not from anything in the Spine."
            ),
        ).alignment = Alignment(wrap_text=True, vertical="top")
        return

    for row_index, entry in enumerate(entries, start=2):
        sheet.cell(row=row_index, column=1, value=entry).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
